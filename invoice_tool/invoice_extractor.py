#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票识别统计工具 v1.4
支持压缩包（ZIP/RAR/7Z）和 PDF/JPG 输入
使用 AI 视觉模型 / 文本模型提取发票信息，输出 Excel
支持多页PDF自动识别（独立小票/合并发票）
可滚动布局，自适应任意屏幕
"""
# ruff: noqa: F821, F841

import os
import sys
import re
import json
import base64
import shutil
import tempfile
import zipfile
import tarfile
import subprocess
import configparser
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import time

import fitz  # PyMuPDF
from PIL import Image
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  配置文件路径（存在用户目录下，持久保存模型设置）
# ─────────────────────────────────────────────
CONFIG_PATH = Path.home() / ".invoice_tool_config.ini"

# 视觉模型（图片识别，默认 GLM-4.6V）
DEFAULT_VISION_URL   = "http://172.18.100.141:8080/v1"
DEFAULT_VISION_MODEL = "GLM-4.6V"
# 文本模型（PDF有文本层时直接用文本，默认 Qwen3.5，API Key 待用户配置）
DEFAULT_TEXT_URL     = ""
DEFAULT_TEXT_MODEL   = "Qwen3.5-122B-A10B"

def load_config() -> dict:
    cfg = configparser.ConfigParser()
    if CONFIG_PATH.exists():
        cfg.read(str(CONFIG_PATH), encoding="utf-8")
    sec = cfg["model"] if "model" in cfg else {}
    return {
        "vision_url":   sec.get("vision_url",   DEFAULT_VISION_URL),
        "vision_model":  sec.get("vision_model",  DEFAULT_VISION_MODEL),
        "vision_key":    sec.get("vision_key",    ""),
        "text_url":      sec.get("text_url",      DEFAULT_TEXT_URL),
        "text_model":    sec.get("text_model",    DEFAULT_TEXT_MODEL),
        "text_key":      sec.get("text_key",      ""),
    }

def save_config(vision_url, vision_model, vision_key, text_url, text_model, text_key):
    cfg = configparser.ConfigParser()
    cfg["model"] = {
        "vision_url":   vision_url,
        "vision_model":  vision_model,
        "vision_key":    vision_key,
        "text_url":      text_url,
        "text_model":    text_model,
        "text_key":      text_key,
    }
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        cfg.write(f)

# ─────────────────────────────────────────────
#  Prompt
# ─────────────────────────────────────────────
PROMPT = """请仔细分析这张发票/收据/单据图片，尽可能提取以下字段信息并以JSON格式输出。
本批次发票可能来自越南、中国等多个国家，可能含有越南语、中文、英文，请全力识别。

【重要】多页PDF的处理方式：
1. 如果每页是独立的"小票/收据照片"（内容独立、金额独立），则把每页当一条记录，用 page_1、page_2 等作为 key：
   {"page_1": {"buyer": "...", "seller": "...", "total_amount": "数字", ...}, "page_2": {...}}
2. 如果多页是"同一个发票的不同部分"（多页合成一张发票），则只提取"总金额"字段（如 Tổng thanh toán / Total payment），忽略单页小计：
   {"buyer": "...", "seller": "...", "total_amount": "总金额数字", "currency": "VND", "date": "...", "invoice_number": "..."}

提取规则：
- 如果某个字段确实无法识别，将值设为空字符串 ""
- 绝对不要虚构或猜测任何信息
- total_amount 只填数字，不含货币符号（如 1234.56）
- date 格式为 YYYY-MM-DD，越南发票日期格式常见 ngày DD tháng MM năm YYYY
- 越南发票：Người mua hàng/Đơn vị = 购方；Người bán hàng/Công ty bán = 销方
- 越南语总金额关键词：Tổng cộng / Tổng tiền thanh toán / Thành tiền / Tổng thanh toán (Total payment)
- 发票号码：Số hoá đơn / Số / Invoice No
- 币种：越南盾填 VND，人民币填 CNY，美元填 USD

需要提取的字段：
- buyer: 购方名称（购买方/付款方/Người mua）
- seller: 销方名称（销售方/开票方/Người bán/Công ty）
- description: 发票说明/商品或服务描述（简要概括，越南语可保留原文）
- total_amount: 总金额（纯数字，多页发票只填汇总金额，不要填单页小计）
- currency: 币种
- date: 开票日期（YYYY-MM-DD）
- invoice_number: 发票号码/单据号

只输出JSON，不要任何额外说明。单页发票示例：
{
  "buyer": "...",
  "seller": "...",
  "description": "...",
  "total_amount": "...",
  "currency": "...",
  "date": "...",
  "invoice_number": "..."
}
多页独立小票示例：
{
  "page_1": {"buyer": "...", "seller": "...", "description": "...", "total_amount": "123", "currency": "VND", "date": "2026-04-02", "invoice_number": "..."},
  "page_2": {"buyer": "...", "seller": "...", "description": "...", "total_amount": "456", "currency": "VND", "date": "2026-04-02", "invoice_number": "..."}
}"""

# ─────────────────────────────────────────────
#  文件提取
# ─────────────────────────────────────────────

# 所有支持的压缩包格式（用于递归解压判断）
ARCHIVE_EXTS = {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz", ".tgz"}


def pdf_has_text(pdf_path: str, min_chars: int = 50) -> bool:
    """判断 PDF 是否有实质文本层"""
    try:
        doc = fitz.open(pdf_path)
        total = 0
        for page in doc:
            total += len(page.get_text().strip())
        doc.close()
        return total >= min_chars
    except Exception:
        return False


def _safe_filename(orig_name: str) -> str:
    for enc in ("gbk", "utf-8", "latin-1", "cp437"):
        try:
            return orig_name.encode(enc).decode("utf-8", errors="strict")
        except (UnicodeDecodeError, LookupError):
            try:
                return orig_name.encode("cp437").decode(enc)
            except (UnicodeDecodeError, LookupError):
                pass
    return orig_name.encode("utf-8", errors="replace").decode("utf-8", errors="replace")


def _extract_zip(archive_path: Path, dest_dir: Path) -> bool:
    try:
        with zipfile.ZipFile(archive_path, "r") as zf:
            for info in zf.infolist():
                safe_name = _safe_filename(info.filename)
                safe_name = safe_name.lstrip("/").replace("..", "")
                info.filename = safe_name
                zf.extract(info, dest_dir)
        return True
    except Exception as e:
        print(f"ZIP 解压失败: {e}")
        return False


def _extract_tar(archive_path: Path, dest_dir: Path) -> bool:
    try:
        with tarfile.open(archive_path, "r:*") as tf:
            tf.extractall(dest_dir)
        return True
    except Exception as e:
        print(f"tar 解压失败: {e}")
        return False


def _find_bundled_7z() -> str | None:
    """在 PyInstaller 打包目录里找 7z.exe"""
    import sys
    if not getattr(sys, "frozen", False):
        return None
    base = getattr(sys, "_MEIPASS", "")
    if not base:
        return None
    # 7z.exe 放在 _MEIPASS/7z/ 目录下
    bundled = Path(base) / "7z" / "7z.exe"
    return str(bundled) if bundled.exists() else None


def _extract_external(archive_path: Path, dest_dir: Path) -> bool:
    # ① 优先系统自带命令
    for cmd in [
        ["7z",    "x", str(archive_path), f"-o{dest_dir}", "-y"],
        ["bsdtar", "xf", str(archive_path), "-C", str(dest_dir)],
        ["unar",  "-o", str(dest_dir), str(archive_path)],
    ]:
        if shutil.which(cmd[0]):
            result = subprocess.run(cmd, capture_output=True)
            if result.returncode == 0:
                return True

    # ② 找不到系统命令 → 尝试打包的 7z.exe
    bundled_7z = _find_bundled_7z()
    if bundled_7z:
        result = subprocess.run(
            [bundled_7z, "x", str(archive_path), f"-o{dest_dir}", "-y"],
            capture_output=True
        )
        if result.returncode == 0:
            return True
        # 7z 也失败 → 再试 rarfile（系统有 UnRAR.dll 时）
        try:
            import rarfile
            with rarfile.RarFile(str(archive_path)) as rf:
                rf.extractall(str(dest_dir))
            return True
        except Exception:
            pass

    return False


def _extract_single(archive_path: str, dest_dir: str) -> bool:
    archive_path = Path(archive_path)
    ext = archive_path.suffix.lower()
    if ext == ".zip":
        return _extract_zip(archive_path, Path(dest_dir))
    elif ext in (".tar", ".gz", ".bz2", ".xz", ".tgz"):
        return _extract_tar(archive_path, Path(dest_dir))
    elif ext in (".rar",):
        return _extract_external(archive_path, Path(dest_dir))
    else:
        return _extract_external(archive_path, Path(dest_dir))


def _recursive_extract(root_dir: Path, max_depth: int = 6, depth: int = 0):
    if depth > max_depth:
        return
    found_archives = [
        p for p in root_dir.rglob("*")
        if p.is_file() and p.suffix.lower() in ARCHIVE_EXTS
    ]
    if not found_archives:
        return
    newly_extracted = False
    for arch in found_archives:
        sub_dir = root_dir / arch.stem
        sub_dir.mkdir(exist_ok=True)
        ok = _extract_single(str(arch), str(sub_dir))
        if ok:
            try:
                arch.unlink()
            except Exception:
                pass
            newly_extracted = True
    if newly_extracted:
        _recursive_extract(root_dir, max_depth, depth + 1)


def extract_archive(archive_path: str, dest_dir: str) -> bool:
    ok = _extract_single(archive_path, dest_dir)
    if not ok:
        return False
    _recursive_extract(Path(dest_dir))
    return True


def pdf_to_images(pdf_path: str, dpi: int = 200) -> list:
    images = []
    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)
        doc.close()
    except Exception as e:
        print(f"PDF 转图片失败 {pdf_path}: {e}")
    return images


def image_to_base64(img: Image.Image, max_size: int = 2000) -> str:
    w, h = img.size
    if max(w, h) > max_size:
        ratio = max_size / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
    from io import BytesIO
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=90)
    return base64.b64encode(buf.getvalue()).decode()


def collect_files(input_path: str, work_dir: str) -> list:
    """收集待识别文件，返回 [{"source_file","folder","images":[PIL],"pdf_path":str}, ...]"""
    input_path = Path(input_path)
    result = []

    # 情况1：压缩包 → 解压后扫描解压目录
    if input_path.suffix.lower() in (".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz", ".tgz"):
        extract_dir = Path(work_dir) / "extracted"
        extract_dir.mkdir(exist_ok=True)
        ok = extract_archive(str(input_path), str(extract_dir))
        if not ok:
            raise RuntimeError(f"无法解压文件: {input_path}")
        _collect_from_dir(extract_dir, result)

    # 情况2：单个文件 → 只处理该文件
    elif input_path.is_file():
        images = []
        pdf_path = None
        if input_path.suffix.lower() == ".pdf":
            images = pdf_to_images(str(input_path))
            pdf_path = str(input_path)
        else:
            try:
                images = [Image.open(str(input_path)).convert("RGB")]
            except Exception as e:
                print(f"图片读取失败 {input_path}: {e}")
        if images:
            result.append({
                "source_file": input_path.name,
                "folder": "",
                "images": images,
                "pdf_path": pdf_path,
            })

    # 情况3：目录 → 扫描整个目录
    elif input_path.is_dir():
        _collect_from_dir(input_path, result)

    return result


def _collect_from_dir(root: Path, result: list):
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if path.suffix.lower() not in (".pdf", ".jpg", ".jpeg", ".png", ".bmp", ".tiff"):
            continue
        try:
            rel = path.relative_to(root)
        except ValueError:
            rel = path.name
        folder = path.parent.name if path.parent != root else ""
        images = []
        pdf_path = None
        if path.suffix.lower() == ".pdf":
            images = pdf_to_images(str(path))
            pdf_path = str(path)
        else:
            try:
                img = Image.open(str(path)).convert("RGB")
                images.append(img)
            except Exception as e:
                print(f"图片读取失败 {path}: {e}")
        if images:
            result.append({
                "source_file": str(rel),
                "folder": folder,
                "images": images,
                "pdf_path": pdf_path,
            })


# ─────────────────────────────────────────────
#  AI 识别（视觉模型 + 文本模型，带重试）
# ─────────────────────────────────────────────

def _parse_json_response(text: str) -> list:
    """从 AI 响应文本中提取 JSON，兼容单页和多页格式，返回 list[dict]"""
    if not text:
        return [{}]
    text = text.strip()
    text = re.sub(r"```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    text = text.rstrip("`").strip()
    try:
        obj = json.loads(text)
        if isinstance(obj, dict) and not any(k.startswith("page_") for k in obj):
            return [obj]
        if isinstance(obj, dict):
            pages = []
            for k in sorted(obj.keys(),
                           key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 0):
                if isinstance(obj[k], dict):
                    pages.append(obj[k])
            return pages if pages else [obj]
    except Exception:
        pass
    m = re.search(r"\{[\s\S]*\}", text)
    if m:
        try:
            obj = json.loads(m.group())
            if isinstance(obj, dict) and not any(k.startswith("page_") for k in obj):
                return [obj]
            if isinstance(obj, dict):
                pages = []
                for k in sorted(obj.keys(),
                               key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 0):
                    if isinstance(obj[k], dict):
                        pages.append(obj[k])
                return pages if pages else [obj]
        except Exception:
            pass
    return [{}]


def _call_vision_api(b64: str, model: str, api_key: str,
                     base_url: str, log_fn=None) -> list:
    """调用视觉模型 API，返回解析后的 list"""
    from openai import OpenAI
    clean_url = re.sub(r"/chat/completions$", "", base_url.rstrip("/"))
    client = OpenAI(api_key=api_key or "none", base_url=clean_url)
    resp = client.chat.completions.create(
        model=model,
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url",
                 "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                {"type": "text", "text": PROMPT},
            ],
        }],
    )
    raw = resp.choices[0].message.content or ""
    if log_fn:
        preview = raw[:120].replace("\n", " ")
        log_fn(f"    视觉模型返回: {preview}{'...' if len(raw)>120 else ''}")
    return _parse_json_response(raw)


def call_ai(images: list, vision_url: str, vision_model: str, vision_key: str,
            log_fn=None, max_retries: int = 2) -> list:
    """调用视觉模型识别图片列表，返回 list[dict]"""
    page_results = []
    for page_idx, img in enumerate(images):
        b64 = image_to_base64(img)
        for attempt in range(1, max_retries + 2):
            try:
                parsed = _call_vision_api(b64, vision_model, vision_key, vision_url, log_fn)
                page_results.extend(parsed)
                break
            except Exception as e:
                if log_fn:
                    log_fn(f"    第{page_idx+1}页 第{attempt}次尝试失败: {e}"
                           + ("，重试中..." if attempt <= max_retries else ""))
                if attempt <= max_retries:
                    time.sleep(1.5)
        else:
            page_results.append({})

    if not page_results:
        return [{"buyer": "", "seller": "", "description": "",
                 "total_amount": "", "currency": "", "date": "", "invoice_number": ""}]
    if len(page_results) > 1:
        return page_results
    merged = {"buyer": "", "seller": "", "description": "",
              "total_amount": "", "currency": "", "date": "", "invoice_number": ""}
    for r in page_results:
        for k in merged:
            if not merged[k] and r.get(k):
                merged[k] = str(r[k]).strip()
    return [merged]


def call_ai_text(pdf_path: str, text_url: str, text_model: str,
                 text_key: str, log_fn=None, max_retries: int = 2) -> list:
    """用文本模型直接识别 PDF 文本层，返回 list[dict]"""
    from openai import OpenAI
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text()
    doc.close()
    if not full_text.strip():
        return [{}]

    PROMPT_TEXT = PROMPT + "\n\n以下为从PDF直接提取的文本内容，请从中提取发票字段：\n"
    clean_url = re.sub(r"/chat/completions$", "", text_url.rstrip("/"))
    client = OpenAI(api_key=text_key or "none", base_url=clean_url)
    for attempt in range(1, max_retries + 2):
        try:
            resp = client.chat.completions.create(
                model=text_model,
                max_tokens=1024,
                messages=[{"role": "user",
                           "content": PROMPT_TEXT + full_text[:8000]}],
            )
            raw = resp.choices[0].message.content or ""
            if log_fn:
                preview = raw[:120].replace("\n", " ")
                log_fn(f"    文本模型返回: {preview}{'...' if len(raw)>120 else ''}")
            return _parse_json_response(raw)
        except Exception as e:
            if log_fn:
                log_fn(f"    文本模型第{attempt}次尝试失败: {e}"
                       + ("，重试中..." if attempt <= max_retries else ""))
            if attempt <= max_retries:
                time.sleep(1.5)
    return [{}]


# ─────────────────────────────────────────────
#  Excel 输出
# ─────────────────────────────────────────────

HEADERS = ["序号", "文件夹", "文件名", "购方", "销方", "发票说明", "总金额", "币种", "日期", "发票号码"]
FIELD_MAP = {
    "购方": "buyer", "销方": "seller", "发票说明": "description",
    "总金额": "total_amount", "币种": "currency", "日期": "date", "发票号码": "invoice_number",
}
COL_WIDTHS = [6, 20, 35, 25, 25, 30, 14, 10, 14, 20]

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="微软雅黑", size=11)
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
TOTAL_FILL    = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT    = Font(bold=True, name="微软雅黑", size=11)
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def save_excel(rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票识别结果"

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

    amount_col = HEADERS.index("总金额") + 1
    amount_rows = []

    for i, row in enumerate(rows, start=1):
        r = i + 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=row.get("folder", ""))
        ws.cell(row=r, column=3, value=row.get("source_file", ""))

        for col_idx, header in enumerate(HEADERS[3:], 4):
            field = FIELD_MAP[header]
            val = row.get(field, "")
            cell = ws.cell(row=r, column=col_idx, value=val)
            if header == "总金额" and val:
                try:
                    cell.value = float(str(val).replace(",", "").replace(" ", ""))
                    cell.number_format = '#,##0.00'
                    amount_rows.append(r)
                except ValueError:
                    cell.value = val

        fill = ALT_FILL if i % 2 == 0 else None
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 20

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=1).font = TOTAL_FONT
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=amount_col - 1)

    if amount_rows:
        s = sum(
            ws.cell(row=r, column=amount_col).value
            for r in amount_rows
            if isinstance(ws.cell(row=r, column=amount_col).value, (int, float))
        )
        sum_cell = ws.cell(row=total_row, column=amount_col, value=s)
        sum_cell.number_format = '#,##0.00'
        sum_cell.font = TOTAL_FONT
        sum_cell.alignment = Alignment(horizontal="right", vertical="center")
    else:
        ws.cell(row=total_row, column=amount_col, value="-")

    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = TOTAL_FILL
        cell.border = BORDER
    ws.row_dimensions[total_row].height = 24

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    wb.save(output_path)


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────

class InvoiceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("发票识别统计工具 v1.4")
        self.geometry("840x680")
        self.resizable(True, True)
        self.configure(bg="#F5F7FA")
        cfg = load_config()
        self._cfg = cfg
        self._build_ui()

    # ── UI 构建 ─────────────────────────────
    def _build_ui(self):
        # ── 固定标题区（不随滚动）────────────
        title_frame = tk.Frame(self, bg="#1F4E79", pady=10)
        title_frame.pack(fill=tk.X)
        tk.Label(
            title_frame, text="🧾 发票识别统计工具",
            font=("微软雅黑", 17, "bold"), fg="white", bg="#1F4E79"
        ).pack()
        tk.Label(
            title_frame, text="支持 ZIP/RAR/PDF/JPG → Excel  |  视觉模型 + 文本模型",
            font=("微软雅黑", 9), fg="#A9C4E0", bg="#1F4E79"
        ).pack()

        # ── 可滚动内容区 ──────────────────────
        # Canvas + Scrollbar 实现自适应滚动
        self._canvas = tk.Canvas(self, bg="#F5F7FA",
                                  highlightthickness=0)
        v_scroll = tk.Scrollbar(self, orient="vertical",
                                 command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=v_scroll.set)

        v_scroll.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        # 真正装内容的 Frame（放在 Canvas 里）
        self._body = tk.Frame(self._canvas, bg="#F5F7FA", padx=20, pady=12)
        self._body_win = self._canvas.create_window(
            (0, 0), window=self._body, anchor="nw"
        )

        # Canvas 内容变化 → 更新滚动区域
        self._body.bind("<Configure>", self._on_body_configure)
        self._canvas.bind("<Configure>", self._on_canvas_configure)

        # 鼠标滚轮支持（Windows/macOS）
        self._canvas.bind_all("<MouseWheel>",
                              lambda e: self._canvas.yview_scroll(
                                  int(-1 * (e.delta / 120)), "units"))
        self._canvas.bind("<Button-4>",
                          lambda e: self._canvas.yview_scroll(-3, "units"))
        self._canvas.bind("<Button-5>",
                          lambda e: self._canvas.yview_scroll(3, "units"))

        # ── 视觉模型配置 ──
        self._section(self._body, "🤖 视觉模型（图片识别，用于扫描件/照片）")
        vf = tk.LabelFrame(
            self._body, text="GLM-4.6V 等视觉模型",
            bg="#F5F7FA", bd=1, relief="solid",
            fg="#1F4E79", font=("微软雅黑", 9, "bold"), padx=10, pady=6
        )
        vf.pack(fill=tk.X, pady=(4, 6))

        for (label, var_name, default, width) in [
            ("接口地址", "vision_url_var",  self._cfg.get("vision_url",  ""), 46),
            ("模型名称", "vision_model_var", self._cfg.get("vision_model", "GLM-4.6V"), 28),
        ]:
            row = tk.Frame(vf, bg="#F5F7FA")
            row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=label, width=10, anchor="w",
                     bg="#F5F7FA", font=("微软雅黑", 9)).pack(side=tk.LEFT)
            setattr(self, var_name, tk.StringVar(value=default))
            tk.Entry(row, textvariable=getattr(self, var_name),
                     font=("Consolas", 9), width=width).pack(side=tk.LEFT, ipady=2)

        r3 = tk.Frame(vf, bg="#F5F7FA")
        r3.pack(fill=tk.X, pady=2)
        tk.Label(r3, text="API Key", width=10, anchor="w",
                 bg="#F5F7FA", font=("微软雅黑", 9)).pack(side=tk.LEFT)
        self.vision_key_var = tk.StringVar(value=self._cfg.get("vision_key", ""))
        self.vision_key_entry = tk.Entry(
            r3, textvariable=self.vision_key_var,
            font=("Consolas", 9), width=34, show="*"
        )
        self.vision_key_entry.pack(side=tk.LEFT, ipady=2)
        self.show_vision_key = tk.BooleanVar(value=False)
        tk.Checkbutton(
            r3, text="显示", variable=self.show_vision_key,
            command=self._toggle_vision_key, bg="#F5F7FA", font=("微软雅黑", 8)
        ).pack(side=tk.LEFT, padx=4)

        # ── 文本模型配置 ──
        self._section(self._body, "📝 文本模型（PDF有文本层时直接提取，用于清晰PDF）")
        tf = tk.LabelFrame(
            self._body, text="Qwen3.5 等文本模型",
            bg="#F5F7FA", bd=1, relief="solid",
            fg="#1F4E79", font=("微软雅黑", 9, "bold"), padx=10, pady=6
        )
        tf.pack(fill=tk.X, pady=(4, 6))

        for (label, var_name, default, width) in [
            ("接口地址", "text_url_var",  self._cfg.get("text_url",  ""), 46),
            ("模型名称", "text_model_var", self._cfg.get("text_model", "Qwen3.5-122B-A10B"), 28),
        ]:
            row = tk.Frame(tf, bg="#F5F7FA")
            row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=label, width=10, anchor="w",
                     bg="#F5F7FA", font=("微软雅黑", 9)).pack(side=tk.LEFT)
            setattr(self, var_name, tk.StringVar(value=default))
            tk.Entry(row, textvariable=getattr(self, var_name),
                     font=("Consolas", 9), width=width).pack(side=tk.LEFT, ipady=2)

        r6 = tk.Frame(tf, bg="#F5F7FA")
        r6.pack(fill=tk.X, pady=2)
        tk.Label(r6, text="API Key", width=10, anchor="w",
                 bg="#F5F7FA", font=("微软雅黑", 9)).pack(side=tk.LEFT)
        self.text_key_var = tk.StringVar(value=self._cfg.get("text_key", ""))
        self.text_key_entry = tk.Entry(
            r6, textvariable=self.text_key_var,
            font=("Consolas", 9), width=34, show="*"
        )
        self.text_key_entry.pack(side=tk.LEFT, ipady=2)
        self.show_text_key = tk.BooleanVar(value=False)
        tk.Checkbutton(
            r6, text="显示", variable=self.show_text_key,
            command=self._toggle_text_key, bg="#F5F7FA", font=("微软雅黑", 8)
        ).pack(side=tk.LEFT, padx=4)

        tk.Label(
            tf,
            text="PDF 有文本层时优先使用（更快），失败后自动降级到视觉模型",
            bg="#F5F7FA", fg="#555", font=("微软雅黑", 8)
        ).pack(anchor="w", padx=10, pady=(0, 4))

        # ── 保存按钮 ──
        btn_row = tk.Frame(self._body, bg="#F5F7FA")
        btn_row.pack(fill=tk.X, pady=(4, 8))
        tk.Button(
            btn_row, text="💾 保存模型配置", command=self._save_model_config,
            bg="#2E86C1", fg="white", font=("微软雅黑", 9),
            relief="flat", padx=10, pady=2, cursor="hand2"
        ).pack(side=tk.LEFT)

        # ── 输入文件 ──
        self._section(self._body, "📁 输入文件")
        file_row = tk.Frame(self._body, bg="#F5F7FA")
        file_row.pack(fill=tk.X, pady=(4, 8))
        self.input_var = tk.StringVar()
        tk.Entry(file_row, textvariable=self.input_var,
                 font=("微软雅黑", 10), width=57).pack(side=tk.LEFT, ipady=4)
        tk.Button(
            file_row, text="浏览", command=self._browse_input,
            bg="#2E86C1", fg="white", font=("微软雅黑", 10), relief="flat",
            padx=12, cursor="hand2"
        ).pack(side=tk.LEFT, padx=(8, 0), ipady=4)

        # ── 输出文件 ──
        self._section(self._body, "💾 输出 Excel")
        out_row = tk.Frame(self._body, bg="#F5F7FA")
        out_row.pack(fill=tk.X, pady=(4, 8))
        self.output_var = tk.StringVar(value=str(
            Path.home() / "Desktop" / f"发票识别结果_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        ))
        tk.Entry(out_row, textvariable=self.output_var,
                 font=("微软雅黑", 10), width=57).pack(side=tk.LEFT, ipady=4)
        tk.Button(
            out_row, text="浏览", command=self._browse_output,
            bg="#2E86C1", fg="white", font=("微软雅黑", 10), relief="flat",
            padx=12, cursor="hand2"
        ).pack(side=tk.LEFT, padx=(8, 0), ipady=4)

        # ── 运行按钮 ──
        btn_row2 = tk.Frame(self._body, bg="#F5F7FA")
        btn_row2.pack(fill=tk.X, pady=(6, 4))
        self.run_btn = tk.Button(
            btn_row2, text="▶  开始识别",
            command=self._start,
            bg="#27AE60", fg="white",
            font=("微软雅黑", 13, "bold"),
            relief="flat", padx=30, pady=8, cursor="hand2"
        )
        self.run_btn.pack(side=tk.LEFT)
        self.open_btn = tk.Button(
            btn_row2, text="📂 打开结果",
            command=self._open_result,
            bg="#8E44AD", fg="white",
            font=("微软雅黑", 11), relief="flat",
            padx=16, pady=8, cursor="hand2", state=tk.DISABLED
        )
        self.open_btn.pack(side=tk.LEFT, padx=(12, 0))

        # ── 进度条 ──
        self.progress_var = tk.DoubleVar(value=0)
        self.progress = ttk.Progressbar(
            self._body, variable=self.progress_var, maximum=100,
            length=700, mode="determinate"
        )
        self.progress.pack(fill=tk.X, pady=(8, 2))
        self.progress_label = tk.StringVar(value="就绪")
        tk.Label(self._body, textvariable=self.progress_label,
                 bg="#F5F7FA", fg="#555", font=("微软雅黑", 9)
                 ).pack(anchor="w")

        # ── 日志 ──
        self._section(self._body, "📋 运行日志")
        self.log_box = scrolledtext.ScrolledText(
            self._body, height=7, font=("Consolas", 9),
            bg="#1E1E1E", fg="#D4D4D4", insertbackground="white",
            state=tk.DISABLED, relief="flat"
        )
        self.log_box.pack(fill=tk.BOTH, expand=True, pady=(4, 0))

    def _on_body_configure(self, event=None):
        """内容Frame大小变化 → 更新Canvas滚动区域"""
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas_configure(self, event=None):
        """Canvas宽度变化 → 同步内容Frame宽度（保持响应式）"""
        canvas_w = event.width
        self._canvas.itemconfig(self._body_win, width=canvas_w)

    def _section(self, parent, text):
        tk.Label(
            parent, text=text, bg="#F5F7FA",
            fg="#1F4E79", font=("微软雅黑", 11, "bold")
        ).pack(anchor="w", pady=(4, 0))

    # ── 事件处理 ─────────────────────────────
    def _toggle_vision_key(self):
        self.vision_key_entry.config(show="" if self.show_vision_key.get() else "*")

    def _toggle_text_key(self):
        self.text_key_entry.config(show="" if self.show_text_key.get() else "*")

    def _save_model_config(self):
        save_config(
            self.vision_url_var.get().strip(),
            self.vision_model_var.get().strip(),
            self.vision_key_var.get().strip(),
            self.text_url_var.get().strip(),
            self.text_model_var.get().strip(),
            self.text_key_var.get().strip(),
        )
        messagebox.showinfo("已保存", "模型配置已保存，下次启动自动加载。")

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="选择输入文件",
            filetypes=[
                ("支持的文件", "*.zip *.rar *.7z *.tar *.gz *.pdf *.jpg *.jpeg *.png *.bmp *.tiff"),
                ("所有文件", "*.*"),
            ]
        )
        if path:
            self.input_var.set(path)

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="保存 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if path:
            self.output_var.set(path)

    def _log(self, msg: str):
        self.log_box.config(state=tk.NORMAL)
        self.log_box.insert(tk.END, f"[{datetime.now():%H:%M:%S}] {msg}\n")
        self.log_box.see(tk.END)
        self.log_box.config(state=tk.DISABLED)
        self.update_idletasks()

    def _set_progress(self, val: float, label: str = ""):
        self.progress_var.set(val)
        if label:
            self.progress_label.set(label)
        self.update_idletasks()

    def _start(self):
        input_path  = self.input_var.get().strip()
        output_path = self.output_var.get().strip()
        vision_url   = self.vision_url_var.get().strip()
        vision_model = self.vision_model_var.get().strip()
        vision_key   = self.vision_key_var.get().strip()
        text_url     = self.text_url_var.get().strip()
        text_model   = self.text_model_var.get().strip()
        text_key     = self.text_key_var.get().strip()

        if not input_path:
            messagebox.showerror("错误", "请选择输入文件！"); return
        if not output_path:
            messagebox.showerror("错误", "请设置输出文件路径！"); return
        # 至少配置了一个模型
        has_vision = bool(vision_url and vision_model)
        has_text    = bool(text_url   and text_model)
        if not has_vision and not has_text:
            messagebox.showerror("错误", "请至少配置一个模型（视觉模型 或 文本模型）！"); return

        self.run_btn.config(state=tk.DISABLED)
        self.open_btn.config(state=tk.DISABLED)
        self._set_progress(0, "处理中...")
        self.log_box.config(state=tk.NORMAL)
        self.log_box.delete("1.0", tk.END)
        self.log_box.config(state=tk.DISABLED)

        threading.Thread(
            target=self._run_pipeline,
            args=(input_path, output_path,
                  vision_url, vision_model, vision_key,
                  text_url, text_model, text_key,
                  has_vision, has_text),
            daemon=True
        ).start()

    def _run_pipeline(self, input_path, output_path,
                      vision_url, vision_model, vision_key,
                      text_url, text_model, text_key,
                      has_vision, has_text):
        work_dir = tempfile.mkdtemp(prefix="invoice_")
        rows = []
        try:
            self._log(f"输入: {input_path}")
            self._log(f"视觉模型: {vision_model} @ {vision_url}" if has_vision else "视觉模型: 未配置")
            self._log(f"文本模型: {text_model} @ {text_url}" if has_text else "文本模型: 未配置")
            self._set_progress(5, "正在解压/扫描文件...")

            files = collect_files(input_path, work_dir)
            total = len(files)
            self._log(f"共发现 {total} 个待识别文件")

            if total == 0:
                self._log("⚠️ 未找到任何 PDF/图片文件")
                self._set_progress(100, "完成（无文件）")
                return

            for idx, item in enumerate(files, 1):
                src   = item["source_file"]
                pages = len(item["images"])
                self._log(f"[{idx}/{total}] {src}（{pages} 页）")
                self._set_progress(5 + 90 * idx / total, f"识别第 {idx}/{total} 个文件...")

                try:
                    results = []

                    # 模型选择策略：
                    # ① PDF 有文本层 → 优先文本模型（更快更省），失败则降级视觉
                    # ② 图片 / 扫描 PDF → 直接视觉模型
                    # ③ 两个模型都不可用 → 跳过
                    pdf_file = item.get("pdf_path")
                    is_pdf   = bool(pdf_file)

                    if has_text and is_pdf and pdf_has_text(pdf_file):
                        self._log(f"  → 文本模型（PDF 有文本层）")
                        results = call_ai_text(
                            pdf_file, text_url, text_model, text_key,
                            log_fn=self._log, max_retries=2
                        )
                        has_result = bool(results) and any(
                            results[0].get(k) for k in ["buyer","seller","total_amount"]
                        )
                        if not has_result and has_vision:
                            self._log(f"  → 文本模型未提取到内容，降级到视觉模型")
                            results = call_ai(
                                item["images"], vision_url, vision_model, vision_key,
                                log_fn=self._log, max_retries=2
                            )
                        elif not has_result:
                            self._log(f"  ⚠️ 文本模型未提取内容，且视觉模型未配置")

                    elif has_vision:
                        model_name = vision_model
                        self._log(f"  → 视觉模型（{model_name}）")
                        results = call_ai(
                            item["images"], vision_url, vision_model, vision_key,
                            log_fn=self._log, max_retries=2
                        )

                    else:
                        self._log("  ❌ 无可用模型（请至少配置一个模型）")
                        results = [{}]

                    # 写入结果
                    for i, r in enumerate(results):
                        r["source_file"] = src
                        r["folder"]      = item["folder"]
                        if len(results) > 1:
                            r["source_file"] = f"{src} (第{i+1}页)"
                        rows.append(r)

                    # 打日志
                    for i, r in enumerate(results):
                        filled = sum(1 for k in ["buyer","seller","total_amount","date","invoice_number"]
                                     if r.get(k))
                        status = "✅" if filled >= 3 else "⚠️ 部分字段为空"
                        tag = f"[第{i+1}/{len(results)}页] " if len(results) > 1 else ""
                        self._log(f"  {status} {tag}金额={r.get('total_amount','')}  "
                                  f"日期={r.get('date','')}  发票号={r.get('invoice_number','')}")

                except Exception as e:
                    self._log(f"  ❌ 识别失败: {e}")
                    rows.append({
                        "source_file": src, "folder": item["folder"],
                        "buyer": "", "seller": "", "description": "",
                        "total_amount": "", "currency": "", "date": "", "invoice_number": ""
                    })

            self._log("正在写入 Excel...")
            save_excel(rows, output_path)
            empty_count = sum(1 for r in rows if not any(
                r.get(k) for k in ["buyer","seller","total_amount","invoice_number"]))
            self._set_progress(100, f"✅ 完成！共 {len(rows)} 条，其中 {empty_count} 条未提取到内容")
            self._log(f"✅ 输出: {output_path}")
            if empty_count:
                self._log(f"⚠️ 有 {empty_count} 条未提取到有效内容，可检查图片清晰度或更换模型重试")
            self.after(0, lambda: self.open_btn.config(state=tk.NORMAL))
            messagebox.showinfo(
                "完成",
                f"识别完成！\n共处理 {len(rows)} 条"
                + (f"\n⚠️ 其中 {empty_count} 条未提取到内容" if empty_count else "")
                + f"\n\n结果已保存至:\n{output_path}"
            )
        except Exception as e:
            self._log(f"❌ 出错: {e}")
            self._set_progress(0, "出错，请检查日志")
            messagebox.showerror("错误", str(e))
        finally:
            shutil.rmtree(work_dir, ignore_errors=True)
            self.after(0, lambda: self.run_btn.config(state=tk.NORMAL))

    def _open_result(self):
        path = self.output_var.get().strip()
        if path and Path(path).exists():
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path])
            else:
                subprocess.run(["xdg-open", path])


# ─────────────────────────────────────────────
#  入口
# ─────────────────────────────────────────────
if __name__ == "__main__":
    app = InvoiceApp()
    app.mainloop()
