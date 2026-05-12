#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多页小票提取工具 - GUI版本
支持 PDF 和 图片 文件，自动提取多页/多图中的小票信息
提取字段：购方、销方、发票说明、总金额、币种、日期、发票号码
提取不到则为空，不虚构信息
"""

import os
import sys
import json
import base64
import requests
import traceback
import threading
import platform
from pathlib import Path
from tkinter import (
    Tk, ttk, StringVar, BooleanVar,
    filedialog, messagebox, scrolledtext,
    END, LEFT, BOTH, X, Y, W, E, S, N,
    Frame, Label, Entry, Button, Checkbutton, Listbox, MULTIPLE
)
from datetime import datetime

# PDF处理
try:
    import fitz
except ImportError:
    pass

# Excel处理
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    pass


class MultiPageReceiptExtractorGUI:
    """多页小票提取工具 GUI"""

    def __init__(self, root):
        self.root = root
        self.root.title("多页小票提取工具 v1.0")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)

        # 配置文件路径
        self.config_file = Path(__file__).parent / "config.json"
        self.load_config()

        # 文件列表
        self.file_list = []

        # 创建界面
        self.create_widgets()

        # 绑定拖拽事件
        self.bind_drag_drop()

    def load_config(self):
        """加载配置"""
        if self.config_file.exists():
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
            except Exception:
                self.config = self.get_default_config()
        else:
            self.config = self.get_default_config()
            self.save_config()

    def get_default_config(self):
        """获取默认配置"""
        return {
            "model_api": {
                "base_url": "http://172.18.100.141:8080/v1/chat/completions",
                "api_key": "2b893fe6-fbc8-4c1d-9ea1-714d01526674",
                "model_name": "GLM-4.6V"
            }
        }

    def save_config(self):
        """保存配置"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"保存配置失败: {e}")

    def create_widgets(self):
        """创建界面组件"""
        # 主容器
        main_frame = Frame(self.root, padx=15, pady=15)
        main_frame.pack(fill=BOTH, expand=True)

        # ========== 模型配置区域 ==========
        config_frame = LabelFrame(main_frame, text="模型配置", padx=10, pady=10)
        config_frame.pack(fill=X, pady=(0, 10))

        # API 地址
        row = 0
        Label(config_frame, text="API 地址:").grid(row=row, column=0, sticky=W, pady=5)
        self.entry_api_url = Entry(config_frame, width=60)
        self.entry_api_url.insert(0, self.config.get("model_api", {}).get("base_url", ""))
        self.entry_api_url.grid(row=row, column=1, columnspan=2, sticky=E+W, pady=5, padx=5)

        # API Key
        row += 1
        Label(config_frame, text="API Key:").grid(row=row, column=0, sticky=W, pady=5)
        self.entry_api_key = Entry(config_frame, width=60, show="*")
        self.entry_api_key.insert(0, self.config.get("model_api", {}).get("api_key", ""))
        self.entry_api_key.grid(row=row, column=1, columnspan=2, sticky=E+W, pady=5, padx=5)

        # 模型名称
        row += 1
        Label(config_frame, text="模型名称:").grid(row=row, column=0, sticky=W, pady=5)
        self.entry_model = Entry(config_frame, width=60)
        self.entry_model.insert(0, self.config.get("model_api", {}).get("model_name", ""))
        self.entry_model.grid(row=row, column=1, columnspan=2, sticky=E+W, pady=5, padx=5)

        # 保存配置按钮
        row += 1
        btn_save = Button(config_frame, text="保存配置", command=self.on_save_config)
        btn_save.grid(row=row, column=2, sticky=E, pady=5)

        config_frame.columnconfigure(1, weight=1)

        # ========== 文件选择区域 ==========
        file_frame = LabelFrame(main_frame, text="文件选择（支持 PDF、图片）", padx=10, pady=10)
        file_frame.pack(fill=X, pady=(0, 10))

        # 文件列表框
        list_frame = Frame(file_frame)
        list_frame.grid(row=0, column=0, columnspan=3, sticky=E+W, pady=(0, 5))

        scrollbar = Scrollbar(list_frame)
        self.file_listbox = Listbox(list_frame, height=6, yscrollcommand=scrollbar.set,
                                    font=("Consolas", 10))
        scrollbar.config(command=self.file_listbox.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.file_listbox.pack(side=LEFT, fill=BOTH, expand=True)

        # 按钮行
        btn_row = Frame(file_frame)
        btn_row.grid(row=1, column=0, columnspan=3, sticky=E+W, pady=5)

        btn_add = Button(btn_row, text="添加文件...", command=self.on_add_files)
        btn_add.pack(side=LEFT, padx=(0, 5))

        btn_add_folder = Button(btn_row, text="添加文件夹...", command=self.on_add_folder)
        btn_add_folder.pack(side=LEFT, padx=(0, 5))

        btn_clear = Button(btn_row, text="清空列表", command=self.on_clear_list)
        btn_clear.pack(side=LEFT)

        # 删除选中按钮
        btn_remove = Button(btn_row, text="删除选中", command=self.on_remove_selected)
        btn_remove.pack(side=LEFT, padx=(5, 0))

        # 输出路径
        Label(file_frame, text="输出路径:").grid(row=2, column=0, sticky=W, pady=5)
        self.entry_output = Entry(file_frame)
        self.entry_output.grid(row=2, column=1, sticky=E+W, pady=5, padx=5)
        btn_output = Button(file_frame, text="浏览...", command=self.on_browse_output)
        btn_output.grid(row=2, column=2, sticky=E, pady=5, padx=(5, 0))

        file_frame.columnconfigure(1, weight=1)

        # 拖拽提示
        drag_label = Label(file_frame, text="💡 也可以直接拖拽文件或文件夹到这里", fg="gray")
        drag_label.grid(row=3, column=0, columnspan=3, pady=5)

        # ========== 控制按钮 ==========
        btn_frame = Frame(main_frame)
        btn_frame.pack(fill=X, pady=(0, 10))

        self.btn_start = Button(btn_frame, text="开始提取", command=self.on_start,
                                 bg="#4CAF50", fg="white", font=("Arial", 12, "bold"),
                                 height=2)
        self.btn_start.pack(side=LEFT, padx=(0, 10))

        self.btn_stop = Button(btn_frame, text="停止", command=self.on_stop,
                               state="disabled", bg="#f44336", fg="white",
                               height=2)
        self.btn_stop.pack(side=LEFT)

        # 进度标签
        self.progress_var = StringVar(value="")
        progress_label = Label(btn_frame, textvariable=self.progress_var, font=("Arial", 10))
        progress_label.pack(side=LEFT, padx=20)

        # ========== 日志区域 ==========
        log_frame = LabelFrame(main_frame, text="运行日志", padx=10, pady=10)
        log_frame.pack(fill=BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap=WORD,
                                                   font=("Consolas", 10))
        self.log_text.pack(fill=BOTH, expand=True)

        # ========== 状态栏 ==========
        self.status_var = StringVar(value="就绪")
        self.status_bar = Label(main_frame, textvariable=self.status_var,
                                 bd=1, relief=SUNKEN, anchor=W)
        self.status_bar.pack(fill=X, pady=(10, 0))

    def bind_drag_drop(self):
        """绑定拖拽事件"""
        try:
            from tkinterdnd2 import DND_FILES
            self.file_listbox.drop_target_register(DND_FILES)
            self.file_listbox.dnd_bind('<<Drop>>', self.on_drop_files)
        except ImportError:
            # 不支持拖拽，但不影响使用
            pass

    def on_drop_files(self, event):
        """处理拖拽文件"""
        files = event.data
        # 解析拖拽的文件路径
        import re
        # 处理 macOS 和 Windows 的路径格式
        paths = re.findall(r'/(?:[^/\0]+/)*[^/\0]+', files) if '/' in files else files.split()
        for path in paths:
            path = path.strip()
            if path and os.path.exists(path):
                if os.path.isdir(path):
                    self.add_folder_to_list(path)
                else:
                    self.add_file_to_list(path)

    def add_file_to_list(self, filepath):
        """添加文件到列表"""
        ext = os.path.splitext(filepath)[1].lower()
        if ext in ['.pdf', '.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif']:
            if filepath not in self.file_list:
                self.file_list.append(filepath)
                self.file_listbox.insert(END, filepath)
                # 自动设置输出路径
                if not self.entry_output.get():
                    output_path = os.path.splitext(filepath)[0] + "_提取结果.xlsx"
                    self.entry_output.delete(0, END)
                    self.entry_output.insert(0, output_path)

    def add_folder_to_list(self, folderpath):
        """添加文件夹中的文件到列表"""
        for root, dirs, files in os.walk(folderpath):
            for filename in files:
                filepath = os.path.join(root, filename)
                self.add_file_to_list(filepath)

    def on_add_files(self):
        """添加文件"""
        filenames = filedialog.askopenfilenames(
            title="选择文件",
            filetypes=[
                ("所有支持文件", "*.pdf *.png *.jpg *.jpeg *.bmp *.tiff *.tif"),
                ("PDF 文件", "*.pdf"),
                ("图片文件", "*.png *.jpg *.jpeg *.bmp *.tiff *.tif"),
                ("所有文件", "*.*")
            ]
        )
        for filename in filenames:
            self.add_file_to_list(filename)

    def on_add_folder(self):
        """添加文件夹"""
        folder = filedialog.askdirectory(title="选择文件夹")
        if folder:
            self.add_folder_to_list(folder)

    def on_clear_list(self):
        """清空列表"""
        self.file_list.clear()
        self.file_listbox.delete(0, END)

    def on_remove_selected(self):
        """删除选中的文件"""
        selection = self.file_listbox.curselection()
        for index in reversed(selection):
            self.file_listbox.delete(index)
            del self.file_list[index]

    def on_save_config(self):
        """保存配置"""
        self.config["model_api"]["base_url"] = self.entry_api_url.get().strip()
        self.config["model_api"]["api_key"] = self.entry_api_key.get().strip()
        self.config["model_api"]["model_name"] = self.entry_model.get().strip()
        self.save_config()
        messagebox.showinfo("成功", "配置已保存")

    def on_browse_output(self):
        """选择输出路径"""
        filename = filedialog.asksaveasfilename(
            title="保存结果",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.entry_output.delete(0, END)
            self.entry_output.insert(0, filename)

    def log(self, message, level="INFO"):
        """写日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        color_map = {
            "INFO": "black",
            "SUCCESS": "green",
            "ERROR": "red",
            "WARNING": "orange"
        }
        color = color_map.get(level, "black")

        self.log_text.tag_config(level, foreground=color)
        self.log_text.insert(END, f"[{timestamp}] {message}\n", level)
        self.log_text.see(END)
        self.root.update_idletasks()

    def on_start(self):
        """开始处理"""
        if not self.file_list:
            messagebox.showerror("错误", "请添加文件")
            return

        output_path = self.entry_output.get().strip()
        if not output_path:
            messagebox.showerror("错误", "请设置输出路径")
            return

        # 更新配置
        self.config["model_api"]["base_url"] = self.entry_api_url.get().strip()
        self.config["model_api"]["api_key"] = self.entry_api_key.get().strip()
        self.config["model_api"]["model_name"] = self.entry_model.get().strip()

        # 禁用按钮
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")

        # 在后台线程执行
        self.running = True
        self.thread = threading.Thread(target=self.process_files,
                                        args=(output_path,))
        self.thread.start()

    def on_stop(self):
        """停止处理"""
        self.running = False
        self.log("正在停止...")

    def process_files(self, output_path):
        """处理所有文件"""
        try:
            all_image_paths = []
            file_page_mapping = []  # 记录每个图片对应的原始文件信息

            self.log(f"开始处理 {len(self.file_list)} 个文件...")

            # 收集所有页面/图片
            for file_path in self.file_list:
                if not self.running:
                    self.log("用户停止处理", "WARNING")
                    return

                if not os.path.exists(file_path):
                    self.log(f"文件不存在: {file_path}", "ERROR")
                    continue

                self.log(f"处理文件: {os.path.basename(file_path)}")
                images = self.file_to_images(file_path)
                
                # 记录每张图片的来源信息
                for img_path in images:
                    all_image_paths.append(img_path)
                    file_page_mapping.append({
                        "file": file_path,
                        "image": img_path
                    })

            if not all_image_paths:
                self.log("没有找到可处理的页面", "ERROR")
                return

            self.log(f"共提取 {len(all_image_paths)} 张图片/页面")

            # 提取信息
            results = []
            total = len(all_image_paths)

            for i, img_path in enumerate(all_image_paths, 1):
                if not self.running:
                    self.log("用户停止处理", "WARNING")
                    return

                self.progress_var.set(f"进度: {i}/{total}")
                self.log(f"[{i}/{total}] 正在分析...")
                
                info = self.extract_receipt_info(img_path)
                info["_source_file"] = file_page_mapping[i-1]["file"]
                info["_page_num"] = i
                results.append(info)

                # 显示结果摘要
                amount = info.get("总金额", "")
                currency = info.get("币种", "")
                status = "成功" if info.get("总金额") not in ["", None, "无法识别"] else "部分"
                self.log(f"  {status}: {amount} {currency}", "SUCCESS" if status == "成功" else "WARNING")

                # 清理临时图片
                try:
                    os.remove(img_path)
                except Exception:
                    pass

            # 保存结果
            self.log("正在保存结果...")
            self.save_to_excel(results, output_path)
            self.log(f"结果已保存: {output_path}", "SUCCESS")

            # 统计信息
            valid_amounts = [r.get("总金额") for r in results 
                          if isinstance(r.get("总金额"), (int, float))]
            if valid_amounts:
                total_sum = sum(valid_amounts)
                self.log(f"共提取 {len(results)} 条记录，总金额合计: {total_sum:,.2f}", "SUCCESS")

            # 完成
            self.status_var.set("处理完成")
            self.progress_var.set("")
            self.log("=" * 50, "SUCCESS")
            self.log("处理完成！", "SUCCESS")
            messagebox.showinfo("完成", f"处理完成！\n结果已保存到:\n{output_path}")

        except Exception as e:
            self.log(f"错误: {str(e)}", "ERROR")
            self.log(traceback.format_exc(), "ERROR")
        finally:
            # 恢复按钮
            self.btn_start.config(state="normal")
            self.btn_stop.config(state="disabled")
            self.status_var.set("就绪")
            self.progress_var.set("")

    def file_to_images(self, file_path):
        """将文件转换为图片列表"""
        image_paths = []
        ext = os.path.splitext(file_path)[1].lower()
        temp_dir = Path(__file__).parent / "temp_images"
        temp_dir.mkdir(exist_ok=True)

        try:
            if ext == '.pdf':
                # PDF 转换为图片
                doc = fitz.open(file_path)
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    mat = fitz.Matrix(2, 2)  # 2x 分辨率
                    pix = page.get_pixmap(matrix=mat)
                    img_path = temp_dir / f"page_{page_num + 1}_{os.getpid()}.png"
                    pix.save(str(img_path))
                    image_paths.append(str(img_path))
                doc.close()
            elif ext in ['.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif']:
                # 直接复制图片
                import shutil
                img_path = temp_dir / f"image_{os.getpid()}_{os.path.basename(file_path)}"
                shutil.copy2(file_path, str(img_path))
                image_paths.append(str(img_path))
        except Exception as e:
            self.log(f"文件转换错误: {e}", "ERROR")

        return image_paths

    def extract_receipt_info(self, image_path):
        """提取小票信息"""
        # 编码图片
        try:
            with open(image_path, "rb") as f:
                image_base64 = base64.b64encode(f.read()).decode("utf-8")
        except Exception as e:
            self.log(f"图片读取错误: {e}", "ERROR")
            return self.get_empty_result()

        # 提示词 - 强调不虚构，只输出真实可识别的信息
        prompt = """你是一个专业的发票/收据信息提取助手。请分析这张图片，提取以下信息：

1. 购方 (Buyer/ Purchaser) - 购买方公司或个人名称
2. 销方 (Seller/ Vendor) - 销售方公司名称
3. 发票说明 - 简单描述这是什么类型的收据（如"餐厅发票"、"超市购物小票"、"酒店住宿发票"等）
4. 总金额 (Total Amount) - 小票上的总金额，只提取数字部分
5. 币种 (Currency) - 如 USD, AUD, EUR, CNY, JPY 等
6. 日期 (Date) - 发票日期，格式 YYYY-MM-DD
7. 发票号码 (Invoice Number) - 发票或收据上的编号

【重要】如果某个字段在图片中无法清晰识别，请将该字段留空（不要填任何内容），不要虚构或推测任何信息。

请直接输出JSON格式，不要添加任何说明：
{"购方":"","销方":"","发票说明":"","总金额":"","币种":"","日期":"","发票号码":""}"""

        headers = {
            "Authorization": f"Bearer {self.config['model_api']['api_key']}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": self.config["model_api"]["model_name"],
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_base64}"}},
                    {"type": "text", "text": prompt}
                ]
            }],
            "temperature": 0.1,
            "max_tokens": 2048
        }

        for attempt in range(3):
            try:
                response = requests.post(
                    self.config["model_api"]["base_url"].rstrip('/'),
                    headers=headers,
                    json=payload,
                    timeout=180
                )
                response.raise_for_status()
                result = response.json()
                content = result["choices"][0]["message"]["content"]

                # 去除特殊标记
                content = content.strip()
                if '<|begin_of_box|>' in content:
                    content = content.split('<|begin_of_box|>')[1]
                if '<|end_of_box|>' in content:
                    content = content.split('<|end_of_box|>')[0]

                # 解析 JSON
                info = json.loads(content)

                # 验证并清理字段 - 确保都是字符串，空值或无法识别设为空字符串
                clean_info = {}
                for field in ["购方", "销方", "发票说明", "总金额", "币种", "日期", "发票号码"]:
                    value = info.get(field, "")
                    if value is None or str(value).strip() in ["", "无法识别", "null", "None"]:
                        clean_info[field] = ""
                    else:
                        clean_info[field] = str(value).strip()

                # 转换金额为数字
                amount_str = clean_info.get("总金额", "")
                if amount_str:
                    try:
                        # 移除常见货币符号和分隔符
                        amount_clean = amount_str.replace(",", "").replace(" ", "")
                        # 尝试提取数字部分
                        import re
                        numbers = re.findall(r'[\d.]+', amount_clean)
                        if numbers:
                            clean_info["总金额"] = float(numbers[0])
                        else:
                            clean_info["总金额"] = ""
                    except Exception:
                        clean_info["总金额"] = ""
                else:
                    clean_info["总金额"] = ""

                return clean_info

            except Exception as e:
                if attempt < 2:
                    self.log(f"重试 ({attempt + 1}/3)...")
                    continue
                self.log(f"API错误: {e}", "ERROR")
                return self.get_empty_result()

        return self.get_empty_result()

    def get_empty_result(self):
        """返回空结果"""
        return {
            "购方": "",
            "销方": "",
            "发票说明": "",
            "总金额": "",
            "币种": "",
            "日期": "",
            "发票号码": ""
        }

    def save_to_excel(self, results, output_path):
        """保存为 Excel"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "小票提取结果"

        # 样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # 表头
        headers = ["序号", "来源文件", "购方", "销方", "发票说明", "总金额", "币种", "日期", "发票号码"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 数据
        for row_idx, result in enumerate(results, 2):
            # 序号
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            # 来源文件（只显示文件名）
            source_file = result.get("_source_file", "")
            filename = os.path.basename(source_file) if source_file else ""
            ws.cell(row=row_idx, column=2, value=filename).border = thin_border
            # 其他字段
            ws.cell(row=row_idx, column=3, value=result.get("购方", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=result.get("销方", "")).border = thin_border
            ws.cell(row=row_idx, column=5, value=result.get("发票说明", "")).border = thin_border

            # 总金额
            amt = result.get("总金额", "")
            amt_cell = ws.cell(row=row_idx, column=6, value=amt if amt != "" else "")
            amt_cell.border = thin_border
            if isinstance(amt, (int, float)):
                amt_cell.number_format = '#,##0.00'

            ws.cell(row=row_idx, column=7, value=result.get("币种", "")).border = thin_border
            ws.cell(row=row_idx, column=8, value=result.get("日期", "")).border = thin_border
            ws.cell(row=row_idx, column=9, value=result.get("发票号码", "")).border = thin_border

        # 统计行
        stats_row = len(results) + 2
        ws.cell(row=stats_row, column=1, value="总计").border = thin_border
        for col in range(2, 6):
            ws.cell(row=stats_row, column=col).border = thin_border
            ws.cell(row=stats_row, column=col).fill = total_fill
        
        # 计算总金额
        total = sum(r.get("总金额", 0) for r in results
                   if isinstance(r.get("总金额"), (int, float)))
        
        total_cell = ws.cell(row=stats_row, column=6, value=total if total > 0 else "")
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
        total_cell.fill = total_fill
        if total > 0:
            total_cell.number_format = '#,##0.00'
        
        # 币种汇总
        currencies = set(r.get("币种", "") for r in results if r.get("币种"))
        currency_text = ", ".join(sorted(currencies)) if currencies else ""
        ws.cell(row=stats_row, column=7, value=currency_text).border = thin_border
        ws.cell(row=stats_row, column=7).fill = total_fill
        
        # 记录数
        ws.cell(row=stats_row, column=8, value=f"共 {len(results)} 条记录").border = thin_border
        ws.cell(row=stats_row, column=8).fill = total_fill
        
        for col in range(9, 10):
            ws.cell(row=stats_row, column=col).border = thin_border
            ws.cell(row=stats_row, column=col).fill = total_fill

        # 列宽
        widths = [8, 30, 25, 25, 30, 15, 10, 15, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(output_path)


def main():
    root = Tk()
    app = MultiPageReceiptExtractorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
